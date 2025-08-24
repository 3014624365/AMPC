import cv2
import numpy as np
import os
import math
import openpyxl

def calculate_psnr(img1, img2, max_pixel_value=255):
    img1 = img1.astype(np.float64)
    img2 = img2.astype(np.float64)
    mse = np.mean((img1 - img2) ** 2)
    if mse == 0:
        return float('inf')
    else:
        return 10 * np.log10((max_pixel_value ** 2) / mse)


def write_to_excel(n,file_path, var1, var2, var3, var4, var5):
    # 加载已经存在的Excel文件
    wb = openpyxl.load_workbook(file_path)
    # 选择活动的工作表，或者你可以通过sheet_name参数指定工作表名称
    ws = wb.active

    # 写入变量值到指定列
    ws['A'+str(n)] = var1  # 假设An是第n行的A列
    ws['B'+str(n)] = var2  # 假设Bn是第n行的B列
    ws['C'+str(n)] = var3  # 假设Cn是第n行的C列
    ws['D'+str(n)] = var4  # 假设Dn是第n行的D列
    ws['E'+str(n)] = var5  # 假设En是第n行的E列

    # 保存工作簿
    wb.save(file_path)
    print("Values have been written to the Excel file.")

def getUCIQE(img):
    img_LAB = cv2.cvtColor(img, cv2.COLOR_BGR2LAB)
    img_LAB = np.array(img_LAB, dtype=np.float64)
    coe_Metric = [0.4680, 0.2745, 0.2576]
    img_lum = img_LAB[:, :, 0] / 255.0
    img_a = img_LAB[:, :, 1] / 255.0
    img_b = img_LAB[:, :, 2] / 255.0

    # 色度标准差
    chroma = np.sqrt(np.square(img_a) + np.square(img_b))
    sigma_c = np.std(chroma)

    # 亮度对比度
    img_lum = img_lum.flatten()
    sorted_index = np.argsort(img_lum)
    top_index = sorted_index[int(len(img_lum) * 0.99)]
    bottom_index = sorted_index[int(len(img_lum) * 0.01)]
    con_lum = img_lum[top_index] - img_lum[bottom_index]

    # 饱和度均值
    chroma = chroma.flatten()
    sat = np.divide(chroma, img_lum, out=np.zeros_like(chroma, dtype=np.float64), where=img_lum != 0)
    avg_sat = np.mean(sat)

    uciqe = sigma_c * coe_Metric[0] + con_lum * coe_Metric[1] + avg_sat * coe_Metric[2]
    return uciqe


def uicm(img):
    b, r, g = cv2.split(img)
    RG = r - g
    YB = (r + g) / 2 - b
    m, n, o = np.shape(img)  # img为三维 rbg为二维
    K = m * n
    alpha_L = 0.1
    alpha_R = 0.1  ##参数α 可调
    T_alpha_L = math.ceil(alpha_L * K)  # 向上取整
    T_alpha_R = math.floor(alpha_R * K)  # 向下取整

    RG_list = RG.flatten()
    RG_list = sorted(RG_list)
    sum_RG = 0
    for i in range(T_alpha_L + 1, K - T_alpha_R):
        sum_RG = sum_RG + RG_list[i]
    U_RG = sum_RG / (K - T_alpha_R - T_alpha_L)
    squ_RG = 0
    for i in range(K):
        squ_RG = squ_RG + np.square(RG_list[i] - U_RG)
    sigma2_RG = squ_RG / K

    YB_list = YB.flatten()
    YB_list = sorted(YB_list)
    sum_YB = 0
    for i in range(T_alpha_L + 1, K - T_alpha_R):
        sum_YB = sum_YB + YB_list[i]
    U_YB = sum_YB / (K - T_alpha_R - T_alpha_L)
    squ_YB = 0
    for i in range(K):
        squ_YB = squ_YB + np.square(YB_list[i] - U_YB)
    sigma2_YB = squ_YB / K

    Uicm = -0.0268 * np.sqrt(np.square(U_RG) + np.square(U_YB)) + 0.1586 * np.sqrt(sigma2_RG + sigma2_YB)
    return Uicm


def EME(rbg, L):
    m, n = np.shape(rbg)  # 横向为n列 纵向为m行
    number_m = math.floor(m / L)
    number_n = math.floor(n / L)
    # A1 = np.zeros((L, L))
    m1 = 0
    E = 0
    for i in range(number_m):
        n1 = 0
        for t in range(number_n):
            A1 = rbg[m1:m1 + L, n1:n1 + L]
            rbg_min = np.amin(np.amin(A1))
            rbg_max = np.amax(np.amax(A1))

            if rbg_min > 0:
                rbg_ratio = rbg_max / rbg_min
            else:
                rbg_ratio = rbg_max  ###
            E = E + np.log(rbg_ratio + 1e-5)

            n1 = n1 + L
        m1 = m1 + L
    E_sum = 2 * E / (number_m * number_n)
    return E_sum


def UICONM(rbg, L):  # wrong
    m, n, o = np.shape(rbg)  # 横向为n列 纵向为m行
    number_m = math.floor(m / L)
    number_n = math.floor(n / L)
    A1 = np.zeros((L, L))  # 全0矩阵
    m1 = 0
    logAMEE = 0
    for i in range(number_m):
        n1 = 0
        for t in range(number_n):
            A1 = rbg[m1:m1 + L, n1:n1 + L]
            rbg_min = int(np.amin(np.amin(A1)))
            rbg_max = int(np.amax(np.amax(A1)))
            plip_add = rbg_max + rbg_min - rbg_max * rbg_min / 1026
            if 1026 - rbg_min > 0:
                plip_del = 1026 * (rbg_max - rbg_min) / (1026 - rbg_min)
                if plip_del > 0 and plip_add > 0:
                    local_a = plip_del / plip_add
                    local_b = math.log(plip_del / plip_add)
                    phi = local_a * local_b
                    logAMEE = logAMEE + phi
            n1 = n1 + L
        m1 = m1 + L
    logAMEE = 1026 - 1026 * ((1 - logAMEE / 1026) ** (1 / (number_n * number_m)))
    return logAMEE

def eme(x, window_size):
    """
      Enhancement measure estimation
      x.shape[0] = height
      x.shape[1] = width
    """
    # 计算图像分块的数量
    k1 = x.shape[1]/window_size
    k2 = x.shape[0]/window_size
    # EME公式权重
    w = 2./(k1*k2)
    blocksize_x = window_size
    blocksize_y = window_size
    # 确保图像可以被window_size整除，可以把没法整除多余的像素切掉
    x = x[:blocksize_y*k2, :blocksize_x*k1]
    val = 0
    for l in range(k1):
        for k in range(k2):
            block = x[k*window_size:window_size*(k+1), l*window_size:window_size*(l+1)]
            # 求块区域中灰度最大的值与最小值
            max_ = np.max(block)
            min_ = np.min(block)
            # 合法性检查，避免计算log(0)
            if min_ == 0.0: val += 0
            elif max_ == 0.0: val += 0
            else: val += math.log(max_/min_)
    return w*val


if __name__ == '__main__':
    results=[]
    img_files = []
    #F:\AMPC\AMPC\picture\图片
    path = "F:\\AMPC\\AMPC\\picture\\result\\input"
    sec_path="F:\\AMPC\\AMPC\\picture\\result\\output"
    kp_full = []
    image_extensions = ['.png', '.jpg', '.jpeg', '.bmp', '.gif', '.tiff', '.svg']
    # 遍历目录中的所有文件
    for root, dirs, files in os.walk(path):
        for file in files:
            if any(file.lower().endswith(ext) for ext in image_extensions):
                img_files.append(file)
    n=2
    for name_file in img_files:
        full_path = os.path.join(path, name_file)
        full_sec_path = os.path.join(sec_path, name_file)
        img1 = cv2.imread(full_path)
        img2 = cv2.imread(full_sec_path)
        img1_resized = cv2.resize(img1, (img2.shape[1], img2.shape[0]))
        psnr_value = calculate_psnr(img1_resized, img2)
        uciqe_value=getUCIQE(img2)
        uiqm=uicm(img2)
        r, b, g = cv2.split(img2)
        Uicm = uicm(img2)
        EME_r = EME(r, 8)
        EME_b = EME(b, 8)
        EME_g = EME(g, 8)
        Uism = 0.299 * EME_r + 0.144 * EME_b + 0.587 * EME_g
        Uiconm = UICONM(img2, 8)
        uiqm = 0.0282 * Uicm + 0.2953 * Uism + 3.5753 * Uiconm
        write_to_excel(n,'Answer.xlsx',name_file,'blurry',psnr_value,uciqe_value,uiqm)
        n+=1

